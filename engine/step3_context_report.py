# step3_context_report.py
# 목적: "표현이 N번 나왔다"는 숫자만으로는 작가가 고칠지 말지 판단할 수 없다.
#       그 표현이 들어간 실제 문장을 챕터별로 모아서 보여줘야
#       작가(켄니)가 직접 보고 판단할 수 있다.
#       즉, 코드는 '찾기'만 하고 '판단'은 사람이 한다.
#
# v2: 유의어/대체어 2가지 모드
#   기본(--synonyms 없음): 내장 사전 기반 (무료, 즉시)
#   --synonyms           : Claude API 기반 — 실제 문장 맥락을 보고 3개 제안 (유료, 고품질)

import os
import glob
import re
from collections import defaultdict
from step1_frequency_check import (
    find_repeated_endings, find_connective_patterns,
    find_narration_verbs, find_action_cliches
)

# ── 내장 유의어 사전 (무료 fallback) ────────────────────────────────────────
SYNONYM_MAP = {
    # 접속·대조 구조
    '그러나':        ('허나', '하지만', '그렇지만'),
    '하지만':        ('그러나', '그렇지만', '허나'),
    '그런데':        ('그러자', '한데', '그렇건만'),
    '그리고':        ('그러면서', '아울러', '또한'),
    '한편':          ('이와 함께', '동시에', '반면'),
    '지만':          ('(으)나', '(으)면서도', '(으)되'),
    # 대화 서술동사
    '말했다':        ('전했다', '내뱉었다', '건넸다'),
    '소리쳤다':      ('고함쳤다', '외쳤다', '부르짖었다'),
    '외쳤다':        ('고함쳤다', '부르짖었다', '소리쳤다'),
    '물었다':        ('되물었다', '넌지시 던졌다', '캐물었다'),
    '대답했다':      ('받았다', '응했다', '받아쳤다'),
    '중얼거렸다':    ('웅얼댔다', '혼자 내뱉었다', '되뇌었다'),
    '쏘아붙였다':    ('날카롭게 말했다', '쐐기를 박았다', '매섭게 받아쳤다'),
    '말씀드렸다':    ('아뢰었다', '전해 드렸다', '여쭈었다'),
    # AI 의심 종결어
    '듯했다':        ('보였다', '느껴졌다', '싶었다'),
    '듯이':          ('처럼', '마냥', '양'),
    '것 같았다':     ('보였다', '싶었다', '느껴졌다'),
    '것이었다':      ('셈이었다', '터였다', '노릇이었다'),
    '라고 생각했다': ('싶었다', '마음속에 되뇌었다', '느꼈다'),
    '문득':          ('불현듯', '갑자기', '어느 순간'),
    '어느새':        ('이미', '어느 틈에', '어느 결에'),
    '멍하니':        ('넋을 놓고', '멍하게', '우두커니'),
    '가만히':        ('조용히', '말없이', '잠자코'),
    '그저':          ('다만', '오직', '그냥'),
    '절박':          ('절실한', '간절한', '다급한'),
    '절실':          ('간절한', '애절한', '절박한'),
    '간절':          ('절실한', '애원하듯', '간곡한'),
    '다급':          ('급박한', '절박한', '다급스러운'),
    '공허':          ('허전한', '텅 빈', '허허로운'),
    # 묘사 클리셰
    '텅 빈':         ('비어 있는', '공허한', '허한'),
    '공허한':        ('텅 빈', '허전한', '허허로운'),
    '짙은 어둠':     ('짙게 깔린 어둠', '칠흑 같은 어둠', '어둠이 짙게'),
    '낯설고 이국적': ('생경한', '낯설기만 한', '이질적인'),
    '삭막하고':      ('황량하고', '메마르고', '쓸쓸하고'),
    '무겁고 답답':   ('짓눌리듯', '답답하고 묵직한', '무겁게 짓누르는'),
    '깊고 무거운':   ('묵직하고 깊은', '묘하게 무거운', '깊고 묵직한'),
    '어둡고 적막':   ('칠흑처럼 조용한', '어둠 속에 고요한', '어둡고 고요한'),
}


def get_synonyms_static(phrase):
    """내장 사전에서 유의어 3개를 찾아 반환. 없으면 빈 문자열."""
    if phrase in SYNONYM_MAP:
        return ' / '.join(SYNONYM_MAP[phrase])
    for key, vals in SYNONYM_MAP.items():
        if key in phrase:
            return ' / '.join(vals)
    return ''


# ── API 기반 유의어 생성 ─────────────────────────────────────────────────────

def _load_env():
    from pathlib import Path
    env_path = Path(__file__).parent.parent / '.env'
    if env_path.exists():
        with open(env_path, encoding='utf-8-sig') as f:
            for line in f:
                line = line.strip()
                if '=' in line and not line.startswith('#'):
                    k, v = line.split('=', 1)
                    os.environ.setdefault(k.strip(), v.strip())


def generate_synonyms_api(phrases_with_contexts, model='claude-sonnet-4-6', yes=False):
    """
    Claude API를 사용해 각 표현의 실제 문장을 보고 유의어/대체어 3개를 생성.
    phrases_with_contexts: {표현: [문장1, 문장2, ...]}
    반환: {표현: "대체어1 / 대체어2 / 대체어3"}
    """
    _load_env()
    try:
        import anthropic
    except ImportError:
        print("  [오류] pip install anthropic 필요")
        return {}

    n = len(phrases_with_contexts)
    print(f"\n[API 유의어 생성]")
    print(f"  대상 표현: {n}개")
    print(f"  모델: {model}")
    print(f"  예상 비용: 약 ${n * 0.002:.3f} ~ ${n * 0.005:.3f} (표현당 $0.002~0.005)")

    if not yes:
        ans = input("  진행할까요? (y/n): ").strip().lower()
        if ans != 'y':
            print("  → 취소. 내장 사전으로 대신 제공합니다.")
            return {}

    client = anthropic.Anthropic()
    results = {}

    for i, (phrase, sentences) in enumerate(phrases_with_contexts.items(), 1):
        examples = sentences[:4]
        example_text = '\n'.join(f'  - {s[:120]}' for s in examples)

        prompt = f"""한국어 소설에서 '{phrase}'이(가) 반복 사용되고 있습니다.
아래 실제 문장들의 맥락과 문체를 보고, 이 표현을 대체할 수 있는 유의어/대체어 3개를 제안해주세요.

실제 사용 문장:
{example_text}

요구사항:
- 한국어 소설 문체에 자연스러운 표현
- 원문의 뉘앙스·강도·품사를 유지
- 문장 안에서 바로 교체 가능한 형태
- 설명·번호·이유 없이, 대체어만 '/' 로 구분해 한 줄로 출력

출력 예시: 허나 / 그렇지만 / 하지만"""

        try:
            response = client.messages.create(
                model=model,
                max_tokens=80,
                messages=[{"role": "user", "content": prompt}]
            )
            result = response.content[0].text.strip()
            results[phrase] = result
            print(f"  [{i}/{n}] '{phrase}' → {result}")
        except Exception as e:
            fallback = get_synonyms_static(phrase)
            results[phrase] = fallback if fallback else ''
            print(f"  [{i}/{n}] '{phrase}' API 오류 → 사전 fallback ({e})")

    print(f"\n  API 유의어 생성 완료 ({len(results)}개)")
    return results


# ── 챕터 로딩 ────────────────────────────────────────────────────────────────

def _strip_scene_breaks(text):
    return re.sub(r'_{5,}', ' ', text)


def _load_from_single_file(path):
    from project_utils import split_single_file_chapters
    chapters_list, method = split_single_file_chapters(path)
    if method:
        print(f"  [챕터 구분 방식: {method}]")
    chapters = {}
    for ch_num, title, body in chapters_list:
        name = f'제{ch_num}화.txt' if method == 'pattern' else f'{ch_num:02d}_{title}.txt'
        chapters[name] = _strip_scene_breaks(body)
    return chapters


def short_chapter_label(chapter_name):
    m = re.search(r'제\d+화', chapter_name)
    return m.group(0) if m else chapter_name.replace('.txt', '')


def _load_from_folder(folder):
    chapters = {}
    for path in sorted(glob.glob(os.path.join(folder, '*.txt'))):
        with open(path, encoding='utf-8') as f:
            text = f.read()
        chapters[os.path.basename(path)] = _strip_scene_breaks(text)
    return chapters


def load_chapters(path, mode='auto'):
    if mode == 'auto':
        mode = 'file' if os.path.isfile(path) else 'folder'
    if mode == 'file':
        return _load_from_single_file(path)
    return _load_from_folder(path)


def split_sentences(text):
    text = text.replace('\n', ' ')
    sentences = re.split(r'(?<=[.!?])\s+', text)
    return [s.strip() for s in sentences if s.strip()]


def find_phrase_contexts(chapters, phrase):
    results = defaultdict(list)
    for chapter_name, text in chapters.items():
        sentences = split_sentences(text)
        for s in sentences:
            if phrase in s:
                results[chapter_name].append(s)
    return results


def get_multi_chapter_phrases(chapters, detector_fn, min_chapters=2):
    combined = defaultdict(dict)
    for chapter_name, text in chapters.items():
        for pattern, count in detector_fn(text).items():
            combined[pattern][chapter_name] = count
    return {p: d for p, d in combined.items() if len(d) >= min_chapters}


# ── 리포트 출력 ──────────────────────────────────────────────────────────────

def write_context_report(chapters, phrases, out_path, synonyms=None):
    """
    synonyms: {표현: "대체어1 / 대체어2 / 대체어3"} — None 이면 내장 사전 사용.
    """
    lines = []
    for phrase in sorted(phrases.keys(), key=lambda p: -sum(phrases[p].values())):
        if synonyms is not None:
            syns = synonyms.get(phrase, get_synonyms_static(phrase))
        else:
            syns = get_synonyms_static(phrase)
        syn_str = f'  ▶ 유의어/대체어: {syns}' if syns else ''
        lines.append(
            f"\n## '{phrase}' (총 {sum(phrases[phrase].values())}회, "
            f"{len(phrases[phrase])}개 챕터){syn_str}\n"
        )
        contexts = find_phrase_contexts(chapters, phrase)
        for chapter_name, sentences in contexts.items():
            short_name = short_chapter_label(chapter_name)
            for s in sentences:
                display = s if len(s) <= 120 else s[:120] + '...'
                lines.append(f"- [{short_name}] {display}")
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))


def write_context_excel(chapters, phrases, out_path, synonyms=None):
    """
    열 구성: 표현 / 화수 / 문장 / 유의어·대체어
    synonyms: {표현: "대체어1 / 대체어2 / 대체어3"} — None 이면 내장 사전 사용.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [참고] openpyxl 없음 — Excel 저장 생략 (pip install openpyxl)")
        return

    def _hfont(bold=True, color="000000"):
        return Font(bold=bold, name="맑은 고딕", color=color)
    def _hfill(color):
        return PatternFill("solid", fgColor=color)
    def _thin():
        s = Side(style='thin', color='AAAAAA')
        return Border(left=s, right=s, top=s, bottom=s)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "반복표현목록"

    hdrs = ["표현", "화수", "문장", "유의어/대체어 (3개)"]
    ws.append(hdrs)
    for c in range(1, len(hdrs) + 1):
        cell = ws.cell(1, c)
        cell.font = _hfont(bold=True, color="FFFFFF")
        cell.fill = _hfill("1565C0")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _thin()
    ws.row_dimensions[1].height = 22

    for phrase in sorted(phrases.keys(), key=lambda p: -sum(phrases[p].values())):
        if synonyms is not None:
            syns = synonyms.get(phrase, get_synonyms_static(phrase))
        else:
            syns = get_synonyms_static(phrase)
        contexts = find_phrase_contexts(chapters, phrase)
        for chapter_name, sentences in contexts.items():
            short_name = short_chapter_label(chapter_name)
            for s in sentences:
                display = s if len(s) <= 160 else s[:160] + '...'
                ws.append([phrase, short_name, display, syns])
                r = ws.max_row
                for c in range(1, 5):
                    cell = ws.cell(r, c)
                    cell.border = _thin()
                    cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left')
                    cell.font = Font(name="맑은 고딕", size=9)

    ws.column_dimensions[get_column_letter(1)].width = 18
    ws.column_dimensions[get_column_letter(2)].width = 10
    ws.column_dimensions[get_column_letter(3)].width = 62
    ws.column_dimensions[get_column_letter(4)].width = 28
    ws.freeze_panes = "A2"
    wb.save(out_path)


# ── main ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description="반복 표현 문맥 리포트 생성")
    parser.add_argument('folder', nargs='?', default=None,
                        help="챕터 폴더 (기본: --project의 sample_chapters)")
    parser.add_argument('--project', default=None,
                        help="projects/<이름> 기준으로 기본 경로를 잡는다")
    parser.add_argument('--out', default=None,
                        help="출력 마크다운 경로 (기본: --project 폴더 안 context_report.md)")
    parser.add_argument('--synonyms', action='store_true',
                        help="Claude API로 문맥 기반 유의어 3개 생성 (유료, 고품질)")
    parser.add_argument('--model', default='claude-sonnet-4-6',
                        help="유의어 생성에 사용할 모델 (기본: claude-sonnet-4-6)")
    parser.add_argument('--yes', action='store_true',
                        help="API 비용 확인 없이 바로 실행")
    args = parser.parse_args()

    if args.folder:
        target_folder = args.folder
    elif args.project:
        from project_utils import default_sample_source
        target_folder = default_sample_source(args.project)
    else:
        target_folder = 'sample_chapters'

    if args.out:
        out_path = args.out
    elif args.project:
        from project_utils import project_path
        out_path = project_path(args.project, 'context_report.md')
    else:
        out_path = 'context_report.md'

    chapters = load_chapters(target_folder)

    all_phrases = {}
    for fn in [find_repeated_endings, find_connective_patterns,
               find_narration_verbs, find_action_cliches]:
        all_phrases.update(get_multi_chapter_phrases(chapters, fn))

    # 유의어 생성
    synonyms = None
    if args.synonyms:
        # 각 표현의 예시 문장 수집
        phrases_with_examples = {}
        for phrase in all_phrases:
            contexts = find_phrase_contexts(chapters, phrase)
            sentences = [s for sents in contexts.values() for s in sents]
            phrases_with_examples[phrase] = sentences
        synonyms = generate_synonyms_api(
            phrases_with_examples,
            model=args.model,
            yes=args.yes
        )
        mode_label = f"API ({args.model})"
    else:
        mode_label = "내장 사전"

    write_context_report(chapters, all_phrases, out_path, synonyms=synonyms)
    print(f"리포트 생성 완료: {out_path} ({len(all_phrases)}개 표현, 유의어: {mode_label})")

    excel_path = out_path.replace('.md', '.xlsx')
    if not excel_path.endswith('.xlsx'):
        excel_path += '.xlsx'
    write_context_excel(chapters, all_phrases, excel_path, synonyms=synonyms)
    print(f"Excel 저장 완료: {excel_path}")
