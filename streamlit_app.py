# NOVELDESK — 소설 분석 웹앱 (Streamlit)
# 배포: https://noveldesk-l2xfbasbodfbvaendzic2x.streamlit.app/

import streamlit as st
import sys, os, re, json, io, glob, statistics, tempfile, shutil
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter

BASE_DIR   = Path(__file__).parent
ENGINE_DIR = BASE_DIR / 'engine'
sys.path.insert(0, str(ENGINE_DIR))

# 로컬 실행 여부 — projects 폴더가 있으면 로컬
LOCAL_PROJECTS = BASE_DIR / 'projects'
IS_LOCAL = LOCAL_PROJECTS.exists()

st.set_page_config(
    page_title="NOVELDESK — 소설 분석",
    page_icon="N",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<script>
(function(){
  setInterval(function(){
    try { window.dispatchEvent(new MouseEvent('mousemove')); } catch(e){}
  }, 20000);
  setInterval(function(){
    var txt = (document.body || {}).innerText || '';
    var gone = txt.includes('gone to sleep') || txt.includes('잠들었') ||
               txt.includes('Reconnecting') || txt.includes('Please wait');
    if(gone){ setTimeout(function(){ location.reload(); }, 2000); return; }
    var app = document.querySelector('[data-testid="stApp"]') || document.querySelector('.main');
    if(app && app.innerHTML.trim().length < 50){ setTimeout(function(){ location.reload(); }, 3000); }
  }, 5000);
})();
</script>
<style>
  /* ── 사이드바 ── */
  [data-testid="stSidebar"] { background:#F7F7F5 !important; border-right:1px solid #E0E0DC; }
  [data-testid="stSidebar"] .stButton > button {
    background:#C41E1E !important; color:#fff !important;
    border:none !important; border-radius:6px !important;
    font-weight:700 !important; font-size:.85rem !important;
  }
  [data-testid="stSidebar"] .stButton > button:hover { background:#A01818 !important; }

  /* ── 헤더 ── */
  .nd-header {
    display:flex; align-items:baseline; gap:14px;
    border-bottom:3px solid #C41E1E; padding-bottom:12px; margin-bottom:6px;
  }
  .nd-logo { font-family:Georgia,serif; font-size:1.8rem; letter-spacing:.2em;
             text-transform:uppercase; color:#111; font-weight:400; margin:0; }
  .nd-sub  { font-size:.8rem; color:#777; margin-bottom:22px; margin-top:4px; }

  /* ── STEP 섹션 헤더 ── */
  .step-section-header {
    display:flex; align-items:center; gap:12px;
    background:#F0F0EE; border-left:4px solid #C41E1E;
    padding:10px 16px; border-radius:0 6px 6px 0;
    margin:20px 0 8px 0;
  }
  .step-section-title { font-weight:800; font-size:1rem; color:#111; margin:0; }
  .step-section-desc  { font-size:.78rem; color:#555; margin:2px 0 0 0; }

  .step-section-header-blue  { border-left-color:#1565C0; }
  .step-section-header-green { border-left-color:#2E7D32; }
  .step-section-header-gold  { border-left-color:#E65100; }

  /* ── 배지 ── */
  .badge-free { display:inline-block; background:#E8F5E9; color:#2E7D32;
    font-size:.65rem; font-weight:800; padding:2px 8px; border-radius:20px;
    border:1px solid #C8E6C9; letter-spacing:.03em; vertical-align:middle; }
  .badge-ai   { display:inline-block; background:#FFF3E0; color:#E65100;
    font-size:.65rem; font-weight:800; padding:2px 8px; border-radius:20px;
    border:1px solid #FFE0B2; letter-spacing:.03em; vertical-align:middle; }
  .badge-plus { display:inline-block; background:#FCE4EC; color:#C41E1E;
    font-size:.65rem; font-weight:800; padding:2px 8px; border-radius:20px;
    border:1px solid #F8BBD0; letter-spacing:.03em; vertical-align:middle; }

  /* ── Expander 스타일 ── */
  div[data-testid="stExpander"] {
    border:1px solid #DCDCDA !important; border-radius:8px !important;
    margin-bottom:8px !important; overflow:hidden !important;
  }
  div[data-testid="stExpander"] > details > summary {
    background:#FAFAF8 !important; padding:12px 16px !important;
    font-weight:700 !important;
  }
  div[data-testid="stExpander"] > details > summary:hover { background:#F0F0EE !important; }

  /* ── 정보 박스 ── */
  .info-box {
    background:#EFF6FF; border:1px solid #BFDBFE; border-radius:6px;
    padding:10px 14px; font-size:.82rem; color:#1E40AF; margin-bottom:10px;
  }
  .warn-box {
    background:#FFFBEB; border:1px solid #FCD34D; border-radius:6px;
    padding:10px 14px; font-size:.82rem; color:#92400E; margin-bottom:10px;
  }
  .ok-box {
    background:#F0FDF4; border:1px solid #86EFAC; border-radius:6px;
    padding:10px 14px; font-size:.82rem; color:#166534; margin-bottom:10px;
  }
  /* 구버전 호환 */
  .warn { background:#FFFBEB; border:1px solid #FCD34D; border-radius:6px;
          padding:10px 14px; font-size:.83rem; color:#92400E; }
  .ok   { background:#F0FDF4; border:1px solid #86EFAC; border-radius:6px;
          padding:10px 14px; font-size:.83rem; color:#166534; }

  /* ── 인물 카드 ── */
  .char-card { background:#F5F5F3; border:1px solid #E0E0DC; border-radius:8px;
               padding:12px 14px; text-align:center; }
  .char-name { font-weight:800; font-size:1rem; color:#111; }
  .char-cnt  { font-family:Georgia,serif; font-size:1.5rem; color:#C41E1E; line-height:1.2; }

  /* ── 단계 안내 카드 ── */
  .how-card {
    border-left:3px solid #C41E1E; padding:12px 16px;
    margin-bottom:10px; background:#fff;
    border-radius:0 6px 6px 0; box-shadow:0 1px 3px rgba(0,0,0,.05);
  }
  .how-num   { font-family:Georgia; font-size:1.2rem; color:#C41E1E; font-weight:700; }
  .how-title { font-weight:800; font-size:.95rem; color:#111; margin:2px 0; }
  .how-body  { font-size:.8rem; color:#444; line-height:1.7; }

  /* ── 가이드 배너 ── */
  .genre-banner {
    background:linear-gradient(135deg,#FFF9F9 0%,#FFF 100%);
    border:1px solid #F5CBCB; border-radius:10px;
    padding:16px 20px; margin-bottom:16px;
  }
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# 유틸 함수
# ════════════════════════════════════════════════════════════

def ch_sort(name):
    m = re.search(r'(\d+)', name); return int(m.group(1)) if m else 0

def ch_label(name):
    m = re.search(r'(\d+)', name); return f"{int(m.group(1))}화" if m else name

def split_by_chapter(text):
    pat = re.compile(r'(?:^|\n)(제\s*(\d+)\s*화[^\n]*)', re.MULTILINE)
    matches = list(pat.finditer(text))
    if not matches:
        return {'원고.txt': text.strip()}
    chapters = {}
    for i, m in enumerate(matches):
        num = int(m.group(2))
        start = m.end()
        end   = matches[i+1].start() if i+1 < len(matches) else len(text)
        body  = text[start:end].strip()
        if body:
            chapters[f'제{num:03d}화.txt'] = body
    return chapters

def extract_docx(b):
    from docx import Document
    doc = Document(io.BytesIO(b))
    return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())

def extract_pdf(b):
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(b)) as pdf:
            return '\n'.join(page.extract_text() or '' for page in pdf.pages)
    except Exception:
        return None

def process_files(uploaded_files):
    chapters, errors = {}, []
    source_names = []
    for uf in uploaded_files:
        ext = uf.name.rsplit('.',1)[-1].lower() if '.' in uf.name else ''
        b   = uf.read()
        try:
            if ext == 'txt':
                try:    text = b.decode('utf-8')
                except: text = b.decode('cp949', errors='ignore')
            elif ext == 'docx':
                text = extract_docx(b)
            elif ext == 'pdf':
                text = extract_pdf(b)
                if not text:
                    errors.append(f"{uf.name}: PDF 텍스트 추출 실패"); continue
            elif ext in ('hwp','hwpx'):
                errors.append(f"{uf.name}: HWP → TXT 또는 DOCX로 변환 후 업로드하세요"); continue
            else:
                errors.append(f"{uf.name}: 지원하지 않는 형식"); continue

            # 파일명(확장자 제거) 기록
            base_name = uf.name.rsplit('.',1)[0] if '.' in uf.name else uf.name
            source_names.append(base_name)

            split = split_by_chapter(text)
            if len(split) > 1:
                chapters.update(split)
                st.sidebar.success(f"📄 {uf.name} → {len(split)}개 챕터 자동 분리")
            else:
                key = base_name + '.txt'
                chapters[key] = text.strip()
        except Exception as e:
            errors.append(f"{uf.name}: {e}")

    # 첫 번째 파일명을 기준 소스 파일명으로 저장
    if source_names:
        st.session_state['source_file'] = source_names[0]

    return chapters, errors

def make_report_dir(project_name, source_file=''):
    """로컬 실행 시 분석리포트_<파일명> 폴더 생성. 클라우드에선 None 반환."""
    if not IS_LOCAL:
        return None
    # 파일명 기반 폴더명: 확장자 제거, 특수문자 제거
    base = re.sub(r'[\\/:*?"<>|]', '_', source_file).strip('_') if source_file else ''
    folder_name = f'분석리포트_{base}' if base else f'분석리포트_{datetime.now().strftime("%Y%m%d")}'
    folder = LOCAL_PROJECTS / project_name / folder_name
    folder.mkdir(parents=True, exist_ok=True)
    return folder

def save_local(folder, filename, data_bytes):
    if folder:
        (folder / filename).write_bytes(data_bytes)


# ════════════════════════════════════════════════════════════
# 엔진 함수
# ════════════════════════════════════════════════════════════

def run_step1(chapters):
    from step1_frequency_check import (
        find_repeated_endings, find_connective_patterns,
        find_descriptive_cliches, find_narration_verbs, find_action_cliches
    )
    all_text = '\n'.join(chapters.values())
    return {
        'chapter_count': len(chapters),
        'total_chars':   len(re.sub(r'\s','',all_text)),
        'B': find_repeated_endings(all_text),
        'C': find_connective_patterns(all_text),
        'D': find_descriptive_cliches(all_text),
        'E': find_narration_verbs(all_text),
        'F': find_action_cliches(all_text),
    }

def run_step2_excel(chapters, use_api=False, api_key=None):
    """
    챕터 간 비교 + 컨텍스트 수집 + 유의어 추천 → Excel bytes 반환
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from step1_frequency_check import (
        find_repeated_endings, find_connective_patterns,
        find_descriptive_cliches, find_narration_verbs, find_action_cliches
    )
    from step3_context_report import get_synonyms_static, find_phrase_contexts

    detectors = [
        ('AI 의심 종결어·단어', find_repeated_endings),
        ('접속·대조 구조',      find_connective_patterns),
        ('묘사 클리셰',         find_descriptive_cliches),
        ('대화 서술동사',       find_narration_verbs),
        ('행동 묘사 클리셰',    find_action_cliches),
    ]

    # 챕터 간 비교 — 2개 이상 챕터 반복만 추출
    combined_multi = {}  # pattern → {ch_label: count}
    pattern_category = {}

    for cat_label, fn in detectors:
        ch_data = defaultdict(dict)
        for ch_name, text in chapters.items():
            for p, c in fn(text).items():
                ch_data[p][ch_label(ch_name)] = c
        for p, d in ch_data.items():
            if len(d) >= 2:
                combined_multi[p] = d
                pattern_category[p] = cat_label

    # 컨텍스트 수집 (표현이 들어간 실제 문장)
    all_sentences = []
    for text in chapters.values():
        sents = re.split(r'[.!?。]\s+', text.replace('\n',' '))
        all_sentences.extend(s.strip() for s in sents if s.strip())

    def get_contexts(pattern, max_n=3):
        hits = [s for s in all_sentences if pattern in s]
        return hits[:max_n]

    # 유의어 (API 또는 정적 사전)
    synonyms = {}
    if use_api and api_key:
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=api_key)
            for p in list(combined_multi.keys())[:30]:
                ctxs = get_contexts(p, 3)
                ctx_text = '\n'.join(f'- {s[:100]}' for s in ctxs)
                resp = client.messages.create(
                    model='claude-haiku-4-5-20251001',
                    max_tokens=60,
                    messages=[{"role":"user","content":
                        f"소설에서 '{p}'을 대체할 유의어 3개를 '/'로 구분해서만 출력:\n{ctx_text}"}]
                )
                synonyms[p] = resp.content[0].text.strip()
        except Exception as e:
            st.warning(f"API 유의어 생성 오류: {e} → 내장 사전으로 대체")
            use_api = False

    if not use_api:
        for p in combined_multi:
            synonyms[p] = get_synonyms_static(p)

    # 가나다 순 정렬
    sorted_patterns = sorted(combined_multi.items(), key=lambda x: x[0])

    # Excel 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '반복표현_챕터간비교'

    # 헤더 색상
    header_fill = PatternFill('solid', fgColor='111111')
    header_font = Font(color='FAFAF8', bold=True, size=10)
    red_fill    = PatternFill('solid', fgColor='FFF0F0')
    thin = Border(
        left=Side(style='thin', color='E0E0DC'),
        right=Side(style='thin', color='E0E0DC'),
        top=Side(style='thin', color='E0E0DC'),
        bottom=Side(style='thin', color='E0E0DC'),
    )

    headers = ['표현', '카테고리', '등장 챕터 수', '총 횟수', '챕터별 분포', '실제 문장 예시', '유의어 추천 3가지']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin

    ws.row_dimensions[1].height = 22

    for ri, (pattern, ch_dict) in enumerate(sorted_patterns, 2):
        dist_str = ', '.join(f"{ch}:{cnt}회" for ch, cnt in sorted(ch_dict.items()))
        ctx_str  = ' | '.join(get_contexts(pattern, 2))
        syn_str  = synonyms.get(pattern, '')
        total    = sum(ch_dict.values())
        n_ch     = len(ch_dict)

        row_data = [pattern, pattern_category.get(pattern,''), n_ch, total, dist_str, ctx_str, syn_str]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(ri, ci, val)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin
            if n_ch >= 3:
                cell.fill = red_fill  # 3개 이상 챕터 = 빨강 강조

    # 열 너비
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(sorted_patterns)

def run_step3(chapters, ranks):
    from step5_consistency_check import find_rank_refs, build_rank_pattern
    rank_pattern = build_rank_pattern(ranks)
    rank_order   = {r: i for i, r in enumerate(ranks)}

    timeline = defaultdict(list)
    for ch_name in sorted(chapters.keys(), key=ch_sort):
        ch_num = ch_sort(ch_name)
        for surname, rank, sentence in find_rank_refs(chapters[ch_name], rank_pattern):
            timeline[surname].append((ch_num, rank, sentence))

    regressions = []
    for name, refs in timeline.items():
        prev_idx, prev_ch = None, None
        for ch, rank, sent in refs:
            curr_idx = rank_order.get(rank, -1)
            if prev_idx is not None and curr_idx < prev_idx:
                regressions.append({'name':name,'prev_ch':prev_ch,'prev_rank':ranks[prev_idx],
                                    'curr_ch':ch,'curr_rank':rank,'sentence':sent})
            if curr_idx >= 0:
                prev_idx, prev_ch = curr_idx, ch
    return {'ranks':ranks, 'regressions':regressions, 'timeline':dict(timeline)}

def run_step4_chart(chapters):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm

    for name in ['Malgun Gothic','NanumGothic','AppleGothic','Noto Sans KR']:
        if name in {f.name for f in fm.fontManager.ttflist}:
            plt.rcParams['font.family'] = name; break
    plt.rcParams['axes.unicode_minus'] = False

    from step7_pacing_check import char_count_no_space, sentence_length_stats, dialogue_ratio

    rows = []
    for ch_name in sorted(chapters.keys(), key=ch_sort):
        text = chapters[ch_name]
        chars     = char_count_no_space(text)
        sent_mean, _ = sentence_length_stats(text)
        dia       = dialogue_ratio(text) * 100
        rows.append({'label': ch_label(ch_name), 'chars': chars, 'sent_mean': sent_mean, 'dia': dia})

    if not rows:
        return None, None

    n = len(rows)
    labels = [r['label'] for r in rows]
    x      = list(range(n))

    # A4 가로 비율, 최소 2000px → figsize in inches at 100dpi
    # 가로 최소 20인치(=2000px@100dpi), 챕터 많을수록 더 넓게
    fig_w = max(20, n * 0.55)
    fig_h = fig_w * 0.55   # A4 가로 황금 비율

    fig, axes = plt.subplots(3, 1, figsize=(fig_w, fig_h), sharex=True, dpi=100)
    fig.patch.set_facecolor('#FAFAF8')

    specs = [
        (axes[0], [r['chars']     for r in rows], '#1A1A1A', '글자수 (공백 제외)', '글자수'),
        (axes[1], [r['sent_mean'] for r in rows], '#C41E1E', '문장 평균 길이',      '평균 길이(자)'),
        (axes[2], [r['dia']       for r in rows], '#2A6A3A', '대화 비중 (%)',        '대화 비중(%)'),
    ]

    for ax, vals, color, title, ylabel in specs:
        ax.set_facecolor('#FFFFFF')
        ax.plot(x, vals, marker='o', color=color, linewidth=2, markersize=6)
        avg = statistics.mean(vals)
        ax.axhline(avg, color='#AAAAAA', linestyle='--', linewidth=1)

        # 꼭지점 라벨 — 챕터 수에 따라 폰트 크기 조절
        fs = max(6, min(9, int(140 / n)))
        for i, (v, r) in enumerate(zip(vals, rows)):
            if ylabel == '글자수':
                lbl = f"{r['label']}-{v:,}자"
            elif ylabel == '대화 비중(%)':
                lbl = f"{r['label']}-{v:.0f}%"
            else:
                lbl = f"{r['label']}-{v:.0f}자"
            ax.annotate(lbl, (i, v), textcoords='offset points', xytext=(0,8),
                        ha='center', fontsize=fs, color='#333')

        ax.set_title(title, fontsize=13, fontweight='bold', color='#111', pad=8)
        ax.set_ylabel(ylabel, fontsize=10)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.tick_params(labelsize=9)

    axes[-1].set_xticks(x)
    axes[-1].set_xticklabels(labels, rotation=45, ha='right', fontsize=9)
    axes[-1].set_xlabel('챕터', fontsize=10)
    fig.tight_layout(pad=2.5)

    # PNG bytes
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=100, facecolor='#FAFAF8', bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue(), rows

def run_step5_characters(chapters, char_names_input):
    """사용자가 입력한 인물 이름으로 챕터별 등장 빈도 추적"""
    names = [n.strip() for n in re.split(r'[,，\s]+', char_names_input) if n.strip()]
    if not names:
        return None, []

    # 각 인물의 챕터별 등장 횟수
    data = {}  # name → {ch_label: count}
    for name in names:
        data[name] = {}
        for ch_name in sorted(chapters.keys(), key=ch_sort):
            cnt = chapters[ch_name].count(name)
            if cnt > 0:
                data[name][ch_label(ch_name)] = cnt

    # 공동 등장 챕터 (두 인물이 같은 챕터에 등장)
    all_ch_labels = sorted({ch_label(c) for c in chapters.keys()}, key=lambda x: ch_sort(x+'0'))
    co_appear = defaultdict(int)
    for i, n1 in enumerate(names):
        for n2 in names[i+1:]:
            shared = set(data[n1].keys()) & set(data[n2].keys())
            co_appear[(n1,n2)] = len(shared)

    return data, all_ch_labels, co_appear

def run_step6_ai_suggest(chapters, api_key, model='claude-haiku-4-5-20251001'):
    """STEP 6: Claude API로 반복 표현 진단 + 대안 제안"""
    from step1_frequency_check import (
        find_repeated_endings, find_connective_patterns,
        find_descriptive_cliches, find_narration_verbs
    )
    all_text = '\n'.join(chapters.values())
    patterns = {}
    for fn in [find_repeated_endings, find_connective_patterns, find_descriptive_cliches, find_narration_verbs]:
        for p, c in fn(all_text).items():
            if c >= 3:
                patterns[p] = c

    if not patterns:
        return {}

    # 컨텍스트 수집
    sents = re.split(r'[.!?]\s+', all_text.replace('\n',' '))
    sents = [s.strip() for s in sents if len(s.strip()) > 10]

    import anthropic
    client = anthropic.Anthropic(api_key=api_key)

    results = {}
    items   = sorted(patterns.items(), key=lambda x: -x[1])[:20]
    progress = st.progress(0, text="AI 분석 중...")

    for idx, (p, cnt) in enumerate(items):
        ctx = [s for s in sents if p in s][:3]
        ctx_text = '\n'.join(f'- {s[:120]}' for s in ctx)
        try:
            resp = client.messages.create(
                model=model, max_tokens=150,
                messages=[{"role":"user","content":
                    f"""소설에서 '{p}'이(가) {cnt}회 반복됩니다.
실제 문장:
{ctx_text}

아래 형식으로만 답하세요:
진단: (한 줄)
대안: 대체어1 / 대체어2 / 대체어3"""}]
            )
            results[p] = {'count': cnt, 'response': resp.content[0].text.strip(), 'contexts': ctx}
        except Exception as e:
            results[p] = {'count': cnt, 'response': f'오류: {e}', 'contexts': ctx}
        progress.progress((idx+1)/len(items), text=f"AI 분석 중... ({idx+1}/{len(items)})")

    progress.empty()
    return results


def build_step6_prompt(chapters):
    """STEP 6 무료 프롬프트: 반복 표현 목록 + 진단 요청"""
    from step1_frequency_check import (
        find_repeated_endings, find_connective_patterns,
        find_descriptive_cliches, find_narration_verbs
    )
    all_text = '\n'.join(chapters.values())
    patterns = {}
    for fn in [find_repeated_endings, find_connective_patterns,
                find_descriptive_cliches, find_narration_verbs]:
        for p, c in fn(all_text).items():
            if c >= 3:
                patterns[p] = c
    if not patterns:
        return None, 0
    sents = re.split(r'[.!?]\s+', all_text.replace('\n', ' '))
    sents = [s.strip() for s in sents if len(s.strip()) > 10]
    lines = [
        "소설에서 아래 표현들이 반복적으로 등장합니다. 각 표현에 대해 다음 형식으로 진단해 주세요.\n\n",
        "형식:\n표현: (표현)\n진단: (왜 이 표현이 반복되면 문제인지 1줄)\n대안: 대체어1 / 대체어2 / 대체어3\n\n---\n\n",
        "[반복 표현 목록]\n\n",
    ]
    for p, c in sorted(patterns.items(), key=lambda x: -x[1])[:25]:
        ctx = [s for s in sents if p in s][:2]
        ctx_str = ' / '.join(f'"{s[:80]}"' for s in ctx)
        lines.append(f"• '{p}' — {c}회\n  예시: {ctx_str}\n\n")
    return ''.join(lines), len(patterns)

def _is_single_file(chapters):
    """제X화 구분 없이 단일 파일로 올라온 경우 감지 (ch_sort 값이 모두 9999 초과)"""
    return all(ch_sort(k) > 9999 for k in chapters)

def _single_file_halves(chapters):
    """단일 파일을 전반/후반 텍스트로 반반 분리"""
    full = '\n\n'.join(v for _, v in sorted(chapters.items(), key=lambda x: ch_sort(x[0])))
    mid = len(full) // 2
    return full[:mid], full[mid:]

def build_step7_phase1_prompt(chapters, start, end):
    """STEP 7 1단계 무료 프롬프트: 전반부 떡밥 탐지"""
    if _is_single_file(chapters):
        first_half, _ = _single_file_halves(chapters)
        combined = first_half
        range_label = "전반부(전체 본문 앞 절반)"
    else:
        selected = sorted(
            [(ch_sort(k), v) for k, v in chapters.items() if start <= ch_sort(k) <= end],
            key=lambda x: x[0]
        )
        if not selected:
            return None
        combined = '\n\n'.join(f"=== {n}화 ===\n{text}" for n, text in selected)
        range_label = f"{start}화~{end}화(전반부)"
    return (
        f"당신은 한국 장편소설 전문 편집자입니다. "
        f"아래는 장편소설의 {range_label} 전체 본문입니다.\n\n"
        "이 구간에서 '나중에 회수될 것 같은 설정/암시/약속(떡밥)'을 찾아주세요.\n"
        "떡밥이란: 나중에 다시 다뤄질 것처럼 던져놓은 인물의 비밀, 의미심장한 물건이나 대사, "
        "풀리지 않은 의문, 예고된 사건, 복선이 되는 묘사 등입니다.\n\n"
        "각 떡밥마다 아래 형식으로 작성하세요:\n\n"
        "## 떡밥 N: (한 줄 요약)\n"
        f"- **등장**: {range_label}\n"
        "- **근거 문장**: \"원문 인용\"\n"
        "- **판단 근거**: (1줄)\n"
        "- **회수 예상**: (추측이면 '추측:'으로 시작)\n\n"
        "중요: 실제 본문 문장만 인용하고, 중요도 순으로 정렬하세요.\n\n"
        f"[본문]\n{combined}"
    )

def build_step7_phase2_prompt(foreshadowing_text, chapters, start, end):
    """STEP 7 2단계 무료 프롬프트: 후반부 회수 확인"""
    if _is_single_file(chapters):
        _, second_half = _single_file_halves(chapters)
        combined = second_half
        range_label = "후반부(전체 본문 뒤 절반)"
    else:
        selected = sorted(
            [(ch_sort(k), v) for k, v in chapters.items() if start <= ch_sort(k) <= end],
            key=lambda x: x[0]
        )
        if not selected:
            return None
        combined = '\n\n'.join(f"=== {n}화 ===\n{text}" for n, text in selected)
        range_label = f"{start}화~{end}화(후반부)"
    return (
        "당신은 한국 장편소설 전문 편집자입니다.\n\n"
        "아래 [떡밥 목록]은 전반부 분석에서 찾아낸 복선/설정/암시 목록입니다.\n"
        f"[후반부 본문]은 {range_label} 전체 본문입니다.\n\n"
        "각 떡밥이 후반부에서 실제로 회수됐는지 확인하고 판정하세요.\n\n"
        "판정 기준:\n"
        "- '회수됨': 후반부에 명확히 연결되는 사건/대사가 있다.\n"
        "- '부분 회수': 관련 내용이 나오지만 완전히 해소되지 않았다.\n"
        "- '회수 안 됨': 후반부 전체에서 연결되는 내용이 없다.\n\n"
        "반드시 아래 두 섹션으로 답하세요:\n\n"
        "## 미회수 떡밥 (우선 확인)\n"
        "('회수 안 됨' 항목만. 없으면 '회수 안 된 떡밥 없음')\n\n"
        "## 전체 판정\n"
        "(모든 항목: 판정 + 회수 위치 + 근거 문장)\n\n"
        f"[떡밥 목록]\n{foreshadowing_text}\n\n"
        f"[후반부 본문]\n{combined}"
    )

def build_step8_prompt(chapters, title):
    """STEP 8 무료 프롬프트: 출판사 제출용 자료"""
    sorted_chs = sorted(chapters.items(), key=lambda x: ch_sort(x[0]))
    sample = sorted_chs[:3] + sorted_chs[-2:]
    sample_text = '\n\n'.join(f"=== {k} ===\n{v[:2000]}" for k, v in sample)
    total_chs   = len(chapters)
    total_chars = sum(len(re.sub(r'\s', '', v)) for v in chapters.values())
    return (
        f"당신은 한국 출판사 편집자입니다. "
        f"아래는 장편소설 『{title}』의 챕터 샘플입니다. "
        f"전체 {total_chs}화, 약 {total_chars:,}자 분량입니다.\n\n"
        "아래 형식으로 출판사 제출용 자료를 작성해 주세요:\n\n"
        "## 책 소개글 (400자 이내)\n"
        "(독자의 흥미를 끌 수 있는, 책의 매력을 압축한 소개문)\n\n"
        "## 줄거리 (600자 이내)\n"
        "(전체 이야기 흐름. 결말은 '...으로 이어진다'처럼 열어두세요)\n\n"
        "## 주요 인물 소개\n"
        "(각 인물: 이름, 나이/직업, 성격·역할 2~3줄)\n\n"
        "## 출판사 홍보문구 3가지 (각 30자 이내)\n"
        "(책 표지·광고에 쓸 짧고 임팩트 있는 문구)\n\n"
        f"[샘플 챕터]\n{sample_text}"
    )

_SONNET = "claude-sonnet-4-6"
_PRICE_IN, _PRICE_OUT = 3.0, 15.0   # $ per MTok

def _est_cost(chars, out_tok=4000):
    in_tok = int(chars / 1.7)
    return in_tok/1e6*_PRICE_IN + out_tok/1e6*_PRICE_OUT, in_tok

def run_step7_detect(chapters, api_key, start, end):
    """전반부 떡밥 탐지"""
    import anthropic as ant
    selected = sorted(
        [(ch_sort(k), v) for k, v in chapters.items() if start <= ch_sort(k) <= end],
        key=lambda x: x[0]
    )
    if not selected:
        raise ValueError(f"{start}~{end}화 범위에 챕터가 없습니다.")
    combined = '\n\n'.join(f"=== {n}화 ===\n{text}" for n, text in selected)
    prompt = (
        f"당신은 한국 장편소설 전문 편집자입니다. "
        f"아래는 장편소설의 {start}화부터 {end}화까지(전반부) 전체 본문입니다.\n\n"
        "이 구간에서 '나중에 회수될 것 같은 설정/암시/약속(떡밥)'을 찾아주세요.\n"
        "떡밥이란: 나중에 다시 다뤄질 것처럼 던져놓은 인물의 비밀, 의미심장한 물건이나 대사, "
        "풀리지 않은 의문, 예고된 사건, 복선이 되는 묘사 등을 말합니다.\n\n"
        "각 떡밥마다 아래 형식으로 작성하세요:\n\n"
        "## 떡밥 N: (한 줄 요약)\n"
        f"- **등장**: {start}~{end}화 중 N화\n"
        "- **근거 문장**: \"원문 인용\"\n"
        "- **판단 근거**: (1줄)\n"
        "- **회수 예상**: (1줄, 추측이면 '추측:'으로 시작)\n\n"
        "중요: 실제 본문 문장만 인용하고, 사소하거나 일반적인 묘사는 제외하세요. "
        "떡밥은 중요도 순으로 정렬하세요.\n\n"
        f"[본문]\n{combined}"
    )
    client = ant.Anthropic(api_key=api_key)
    resp = client.messages.create(
        model=_SONNET, max_tokens=4000,
        messages=[{"role": "user", "content": prompt}]
    )
    return resp.content[0].text, resp.usage

def run_step7_resolve(chapters, api_key, foreshadowing_md, start, end):
    """후반부 떡밥 회수 확인"""
    import anthropic as ant
    selected = sorted(
        [(ch_sort(k), v) for k, v in chapters.items() if start <= ch_sort(k) <= end],
        key=lambda x: x[0]
    )
    if not selected:
        raise ValueError(f"{start}~{end}화 범위에 챕터가 없습니다.")
    combined = '\n\n'.join(f"=== {n}화 ===\n{text}" for n, text in selected)
    prompt = (
        "당신은 한국 장편소설 전문 편집자입니다.\n\n"
        "아래 [떡밥 목록]은 전반부를 분석해서 찾아낸 '나중에 회수될 설정/암시/약속' 목록입니다.\n"
        f"[후반부 본문]은 {start}화부터 {end}화까지 전체 본문입니다.\n\n"
        "[떡밥 목록]의 각 항목이 [후반부 본문]에서 실제로 회수됐는지 확인하세요.\n\n"
        "판정 기준:\n"
        "- '회수됨': 후반부에 명확히 연결되는 사건/대사/설명이 있다.\n"
        "- '부분 회수': 관련 내용이 나오지만 완전히 해소되지 않았다.\n"
        "- '회수 안 됨': 후반부 전체에서 연결되는 내용이 없다.\n\n"
        "반드시 아래 두 섹션 형식으로 답하세요.\n\n"
        "## 미회수 떡밥 (우선 확인)\n"
        "('회수 안 됨' 항목만. 없으면 '회수 안 된 떡밥 없음'이라고 쓰세요)\n\n"
        "### 떡밥: (한 줄 요약)\n"
        "- **원래 등장**: (전반부 화수)\n"
        "- **판정**: 회수 안 됨\n"
        "- **검토 의견**: (이유)\n\n"
        "## 전체 판정 (참고)\n"
        "(모든 항목을 빠짐없이 나열)\n\n"
        "### 떡밥: (한 줄 요약)\n"
        "- **원래 등장**: (화수)\n"
        "- **판정**: 회수됨 / 부분 회수 / 회수 안 됨\n"
        "- **회수 위치**: (몇 화 또는 '해당 없음')\n"
        "- **근거 문장**: (원문 인용 또는 생략)\n\n"
        f"[떡밥 목록]\n{foreshadowing_md}\n\n"
        f"[후반부 본문]\n{combined}"
    )
    client = ant.Anthropic(api_key=api_key)
    resp = client.messages.create(
        model=_SONNET, max_tokens=4000,
        messages=[{"role": "user", "content": prompt}]
    )
    return resp.content[0].text, resp.usage

def run_step8_publish(chapters, api_key, title):
    """출판사 제출용 자료 생성"""
    import anthropic as ant
    sorted_chs = sorted(chapters.items(), key=lambda x: ch_sort(x[0]))
    sample = sorted_chs[:3] + sorted_chs[-2:]
    sample_text = '\n\n'.join(f"=== {k} ===\n{v[:2000]}" for k, v in sample)
    total_chs   = len(chapters)
    total_chars = sum(len(re.sub(r'\s', '', v)) for v in chapters.values())
    prompt = (
        f"당신은 한국 출판사 편집자입니다. "
        f"아래는 장편소설 『{title}』의 챕터 샘플입니다. "
        f"전체 {total_chs}화, 약 {total_chars:,}자 분량입니다.\n\n"
        "아래 형식으로 출판사 제출용 자료를 작성해 주세요:\n\n"
        "## 책 소개글 (400자 이내)\n"
        "(독자의 흥미를 끌 수 있는, 책의 매력을 압축한 소개문)\n\n"
        "## 줄거리 (600자 이내)\n"
        "(전체 이야기 흐름. 결말은 '...으로 이어진다'처럼 열어두세요)\n\n"
        "## 주요 인물 소개\n"
        "(각 인물: 이름, 나이/직업, 성격·역할 2~3줄)\n\n"
        "## 출판사 홍보문구 3가지 (각 30자 이내)\n"
        "(책 표지·광고에 쓸 짧고 임팩트 있는 문구)\n\n"
        f"[샘플 챕터]\n{sample_text}"
    )
    client = ant.Anthropic(api_key=api_key)
    resp = client.messages.create(
        model=_SONNET, max_tokens=3000,
        messages=[{"role": "user", "content": prompt}]
    )
    return resp.content[0].text, resp.usage


# ════════════════════════════════════════════════════════════
# 사이드바
# ════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("""
    <div style="font-family:Georgia,serif;font-size:1.1rem;letter-spacing:.2em;
    color:#C41E1E;border-bottom:2px solid #C41E1E;padding-bottom:10px;
    margin-bottom:16px;font-weight:400">NOVELDESK</div>
    """, unsafe_allow_html=True)

    # ── STEP 1: 소설 이름 ──
    st.markdown("**① 소설 이름을 입력하세요**")
    project_name = st.text_input("소설 이름", placeholder="예: 시계이야기", key="proj",
                                 label_visibility="collapsed")

    st.markdown("---")

    # ── STEP 2: 원고 업로드 ──
    st.markdown("**② 원고 파일을 올리세요**")
    st.caption("TXT · DOCX · PDF 지원 | 한 파일에 전권도 OK")
    uploaded = st.file_uploader("파일 선택", type=['txt','docx','pdf'],
                                accept_multiple_files=True, label_visibility="collapsed")

    with st.expander("텍스트 직접 붙여넣기"):
        paste_title = st.text_input("챕터 이름", placeholder="제1화", key="ptitle")
        paste_text  = st.text_area("원고 텍스트", height=100,
                                   placeholder="내용을 붙여넣으세요", key="ptext",
                                   label_visibility="collapsed")
        if st.button("등록", use_container_width=True, key="paste_btn"):
            if paste_text.strip():
                if 'chapters' not in st.session_state:
                    st.session_state.chapters = {}
                t = paste_title.strip() or '원고'
                st.session_state.chapters[f'{t}.txt'] = paste_text.strip()
                st.success(f"'{t}' 등록 완료!")

    st.markdown("---")

    # ── STEP 3: 장르 선택 ──
    st.markdown("**③ 원고 종류를 선택하세요**")
    genre = st.radio("장르", ["소설 (픽션)", "비즈니스/에세이/논픽션"],
                     key="genre", label_visibility="collapsed")
    if genre == "소설 (픽션)":
        st.caption("STEP 1~8 모두 사용 가능합니다")
    else:
        st.caption("STEP 1·2·4·6·8을 사용하세요 (STEP 3·5·7은 소설 전용)")

    st.markdown("---")

    # ── API 키 (접힌 상태) ──
    with st.expander("Claude API 키 입력 (선택)"):
        st.caption("STEP 6+ 자동화에만 필요합니다.\n무료 사용자는 입력 안 해도 됩니다.")
        api_key_input = st.text_input("API Key", type="password",
                                      placeholder="sk-ant-...",
                                      key="api_key", label_visibility="collapsed")
        if api_key_input:
            st.success("API 키 입력됨")

    st.markdown("---")

    # ── 초기화 버튼 ──
    if st.button("새 소설 시작 (전체 초기화)", use_container_width=True):
        for k in list(st.session_state.keys()):
            if k not in ('proj', 'api_key', 'genre'):
                del st.session_state[k]
        st.rerun()

    # ── 지원 형식 안내 ──
    st.markdown("""
    <div style="font-size:.71rem;color:#888;line-height:1.9;margin-top:8px">
    ✅ DOCX (워드)<br>
    ✅ TXT (텍스트)<br>
    ✅ PDF<br>
    ⚠️ HWP → DOCX로 변환 후 업로드
    </div>
    """, unsafe_allow_html=True)

    if IS_LOCAL:
        st.markdown("---")
        st.caption(f"💾 결과 저장: `projects/`")

# ── 파일 처리 ──
if uploaded:
    ch, errs = process_files(uploaded)
    if ch:
        if 'chapters' not in st.session_state:
            st.session_state.chapters = {}
        st.session_state.chapters.update(ch)
    for e in errs:
        st.sidebar.error(e)

chapters  = st.session_state.get('chapters', {})
proj_name = (project_name or 'untitled').strip()

# ════════════════════════════════════════════════════════════
# 메인
# ════════════════════════════════════════════════════════════

st.markdown("""
<div class="nd-header">
  <div class="nd-logo">NOVELDESK</div>
</div>
<div class="nd-sub">AI 소설 분석 도구 &nbsp;·&nbsp; Claude Pro 구독만 있으면 API 키 없이 전체 기능 사용 가능</div>
""", unsafe_allow_html=True)

genre = st.session_state.get('genre', '소설 (픽션)')

# ── 원고 없을 때 안내 ──
if not chapters:
    st.markdown("---")

    # 장르별 추천 STEP 배너
    if genre == "소설 (픽션)":
        st.markdown("""
        <div class="genre-banner">
          <div style="font-weight:800;font-size:.95rem;color:#C41E1E;margin-bottom:6px">장편소설 분석 흐름</div>
          <div style="font-size:.82rem;color:#444;line-height:2">
          STEP 1 반복 표현 탐지 &nbsp;→&nbsp;
          STEP 2 Excel 저장 &nbsp;→&nbsp;
          STEP 4 페이싱 차트 &nbsp;→&nbsp;
          STEP 6 AI 진단 &nbsp;→&nbsp;
          STEP 7 복선 확인 &nbsp;→&nbsp;
          STEP 8 출판 자료
          </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="genre-banner">
          <div style="font-weight:800;font-size:.95rem;color:#E65100;margin-bottom:6px">비소설·논픽션 분석 흐름</div>
          <div style="font-size:.82rem;color:#444;line-height:2">
          STEP 1 반복 표현 탐지 &nbsp;→&nbsp;
          STEP 2 Excel 저장 &nbsp;→&nbsp;
          STEP 4 페이싱 차트 &nbsp;→&nbsp;
          STEP 6 AI 진단 &nbsp;→&nbsp;
          STEP 8 출판 자료
          </div>
          <div style="font-size:.76rem;color:#888;margin-top:6px">STEP 3·5·7은 소설 전용이므로 건너뜁니다</div>
        </div>
        """, unsafe_allow_html=True)

    col1, col2 = st.columns([1, 1])
    with col1:
        st.markdown("### 시작 방법 (3단계)")
        for num, title, body in [
            ("①", "소설 이름 입력", "왼쪽 사이드바 맨 위에 소설 제목을 입력하세요."),
            ("②", "원고 파일 업로드",
             "TXT, DOCX, PDF 파일을 올리세요.<br>"
             "파일 하나에 전권이 들어있어도 됩니다.<br>"
             "<b>'제1화', '제2화'...</b> 패턴이 있으면 챕터를 자동으로 나눕니다."),
            ("③", "STEP 버튼 클릭",
             "파일이 올라가면 아래에 STEP들이 나타납니다.<br>"
             "각 STEP 제목을 클릭해서 펼친 뒤 버튼을 누르세요."),
        ]:
            st.markdown(f"""
            <div class="how-card">
              <div class="how-num">{num}</div>
              <div class="how-title">{title}</div>
              <div class="how-body">{body}</div>
            </div>""", unsafe_allow_html=True)

    with col2:
        st.markdown("### 각 STEP이 하는 일")
        items = [
            ("STEP 1", "FREE", "#2E7D32", "반복 표현 탐지",
             "책 전체에서 AI 의심 패턴, 반복 종결어, 묘사 클리셰를 자동으로 찾아냅니다."),
            ("STEP 2", "FREE", "#2E7D32", "Excel 저장",
             "챕터별 반복 표현을 비교해서 Excel 파일로 저장합니다."),
            ("STEP 3", "FREE", "#2E7D32", "계급·직급 체크 (소설 전용)",
             "군계급·직장직급이 순서 역행하는지 확인합니다."),
            ("STEP 4", "FREE", "#2E7D32", "페이싱 차트",
             "챕터별 글자수·대화 비중을 그래프로 시각화합니다."),
            ("STEP 5", "FREE", "#2E7D32", "인물 등장 추적 (소설 전용)",
             "인물별 챕터 등장 횟수와 공동 등장 관계를 보여줍니다."),
            ("STEP 6", "무료 AI", "#E65100", "AI 반복 표현 진단",
             "Claude.ai에 붙여넣으면 진단 + 대안 표현을 제안합니다. API 키 불필요."),
            ("STEP 7", "무료 AI", "#E65100", "복선 탐지 (소설 전용)",
             "전반부 복선을 찾고 후반부에서 회수됐는지 확인합니다."),
            ("STEP 8", "무료 AI", "#E65100", "출판 자료 생성",
             "소개글·줄거리·홍보문구를 자동으로 만들어줍니다."),
        ]
        for step, badge, badge_color, name, desc in items:
            # 비소설이면 소설 전용 STEP 흐리게
            is_novel_only = "소설 전용" in name
            opacity = "0.4" if (genre != "소설 (픽션)" and is_novel_only) else "1"
            badge_bg = "#E8F5E9" if badge == "FREE" else "#FFF3E0"
            st.markdown(f"""
            <div style="display:flex;gap:10px;align-items:flex-start;
                        margin-bottom:7px;padding:8px 12px;background:#fff;
                        border:1px solid #E0E0DC;border-radius:6px;opacity:{opacity}">
              <div style="font-family:Georgia;font-size:.78rem;font-weight:700;
                          color:#888;min-width:52px;padding-top:2px">{step}</div>
              <div>
                <span style="font-weight:800;font-size:.85rem;color:#111">{name}</span>
                <span style="font-size:.62rem;font-weight:800;padding:2px 7px;
                             border-radius:20px;background:{badge_bg};color:{badge_color};
                             margin-left:6px;border:1px solid {badge_bg}">{badge}</span>
                <div style="font-size:.75rem;color:#666;line-height:1.55;margin-top:3px">{desc}</div>
              </div>
            </div>""", unsafe_allow_html=True)

    st.markdown("---")
    with st.expander("HWP 파일은 어떻게 올리나요?"):
        st.markdown("""
        한글(HWP) 프로그램을 열고:
        1. **파일 → 다른 이름으로 저장**
        2. 파일 형식: **MS 워드 (*.docx)** 선택
        3. 저장한 DOCX 파일을 NOVELDESK에 업로드
        """)
    st.stop()

# ── 현황 요약 ──
sorted_chapters = sorted(chapters.keys(), key=ch_sort)
total_chars = sum(len(re.sub(r'\s','',t)) for t in chapters.values())
src_file   = st.session_state.get('source_file', proj_name)
report_dir = make_report_dir(proj_name, src_file) if IS_LOCAL else None

genre = st.session_state.get('genre', '소설 (픽션)')
is_novel = (genre == "소설 (픽션)")

# ── 현황 요약 ──
c1, c2, c3 = st.columns(3)
c1.metric("소설", proj_name)
c2.metric("전체 글자수", f"{total_chars:,}자")
c3.metric("챕터 수", f"{len(chapters)}개")

ch_preview = ', '.join(ch_label(c) for c in sorted_chapters[:12])
if len(chapters) > 12:
    ch_preview += f" 외 {len(chapters)-12}개"
st.caption(f"등록 챕터: {ch_preview}")

if IS_LOCAL and report_dir:
    st.caption(f"💾 분석 결과 저장 폴더: `분석리포트_{src_file}/`")

# ── 장르별 추천 흐름 배너 ──
if is_novel:
    st.markdown("""
    <div style="background:#FFF9F9;border:1px solid #F5CBCB;border-radius:8px;
    padding:10px 16px;font-size:.8rem;color:#444;margin:12px 0">
    <b style="color:#C41E1E">소설 분석 순서:</b> &nbsp;
    STEP 1 → STEP 2 → STEP 4 → STEP 6 → STEP 7 → STEP 8
    </div>
    """, unsafe_allow_html=True)
else:
    st.markdown("""
    <div style="background:#FFF8F0;border:1px solid #F5D9C0;border-radius:8px;
    padding:10px 16px;font-size:.8rem;color:#444;margin:12px 0">
    <b style="color:#E65100">비소설 분석 순서:</b> &nbsp;
    STEP 1 → STEP 2 → STEP 4 → STEP 6 → STEP 8 &nbsp;
    <span style="color:#999">(STEP 3·5·7은 소설 전용)</span>
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")
st.markdown("""
<div class="step-section-header">
  <div>
    <div class="step-section-title">STEP 1 – 5 &nbsp; 무료 분석</div>
    <div class="step-section-desc">API 키 없이 즉시 사용 가능 &nbsp;·&nbsp; 결과 자동 저장</div>
  </div>
</div>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# STEP 1 — 반복 표현 탐지 (책 전체)
# ════════════════════════════════════════════════════════════

with st.expander("STEP 1 — 반복 표현 탐지  |  무료 · API 불필요"):
    st.markdown("""
    책 전체 텍스트를 **하나로 합쳐서** 반복 표현을 탐지합니다.
    - **[B]** AI 의심 종결어: `듯했다`, `것 같았다`, `문득` 등
    - **[C]** 접속·대조 구조: `지만`, `그러나`, `하지만` 등
    - **[D]** 묘사 클리셰: `텅 빈`, `공허한` 등
    - **[E]** 대화 서술동사: `말했다`, `외쳤다` 등
    - **[F]** 행동 묘사 클리셰: `발길을 재촉했다` 류

    > **STEP 1 vs STEP 2 차이:** STEP 1은 "책 전체에서 몇 번"을 셉니다.
    > STEP 2는 "어느 챕터에서 몇 번씩"인지를 비교해서
    > **2개 이상 챕터에 걸쳐 반복되는 것만** 걸러냅니다 — 이것이 진짜 습관적 패턴입니다.
    """)

    if st.button("STEP 1 실행", key="r1", type="primary"):
        with st.spinner("분석 중..."):
            try:
                r = run_step1(chapters)
                st.session_state.s1 = r
                st.success(f"완료 — {r['chapter_count']}개 챕터 · {r['total_chars']:,}자")
            except Exception as e:
                st.error(f"오류: {e}")

    if 'S1' in {k.upper() for k in st.session_state} and 's1' in st.session_state:
        import pandas as pd
        r = st.session_state.s1
        cols = st.columns(3)
        cols[0].metric("챕터", r['chapter_count'])
        cols[1].metric("글자수", f"{r['total_chars']:,}")
        total_found = sum(len(v) for k,v in r.items() if k in 'BCDEF')
        cols[2].metric("발견 패턴", f"{total_found}종")

        cat_labels = {'B':'[B] AI 의심 종결어','C':'[C] 접속·대조','D':'[D] 묘사 클리셰',
                      'E':'[E] 대화 서술동사','F':'[F] 행동 클리셰'}
        all_rows = []
        for k, lbl in cat_labels.items():
            items = sorted(r.get(k,{}).items(), key=lambda x:-x[1])
            for p, c in items:
                all_rows.append({'카테고리':lbl,'표현':p,'횟수':c})
        if all_rows:
            df = pd.DataFrame(all_rows)
            st.dataframe(df, use_container_width=True, hide_index=True)

            # 저장
            json_bytes = json.dumps(r, ensure_ascii=False, indent=2).encode('utf-8')
            st.download_button("⬇ STEP1 결과 JSON 저장", json_bytes,
                               f"STEP1_반복표현_{src_file}.json", "application/json")
            save_local(report_dir, f"STEP1_반복표현_{src_file}.json", json_bytes)


# ════════════════════════════════════════════════════════════
# STEP 2 — 챕터 간 비교 + Excel (가나다 순 + 유의어)
# ════════════════════════════════════════════════════════════

with st.expander("STEP 2 — 챕터별 반복 표현 비교 + Excel 저장  |  무료"):
    st.markdown("""
    **챕터별로 따로 분석**해서 2개 이상 챕터에 반복되는 표현만 추출합니다.
    결과를 **가나다 순**으로 정렬하고 **유의어 3개**를 함께 Excel로 저장합니다.

    Excel 열 구성: `표현 | 카테고리 | 등장챕터수 | 총횟수 | 챕터별분포 | 실제문장 | 유의어추천`

    - **무료 유의어**: 내장 사전 (즉시, API 불필요)
    - **AI 유의어**: 사이드바에 API 키 입력 → 체크박스 활성화
    """)

    use_api_synonyms = False
    if st.session_state.get('api_key'):
        use_api_synonyms = st.checkbox("Claude AI 유의어 사용 (API 키 필요, 소량 비용 발생)", value=False)

    if st.button("STEP 2 실행 + Excel 생성", key="r2", type="primary"):
        with st.spinner("챕터 간 비교 중 + Excel 생성 중..."):
            try:
                xlsx_bytes, n = run_step2_excel(
                    chapters,
                    use_api=use_api_synonyms,
                    api_key=st.session_state.get('api_key','')
                )
                st.session_state.s2_xlsx = xlsx_bytes
                st.session_state.s2_count = n
                save_local(report_dir, f"STEP2_반복표현Excel_{src_file}.xlsx", xlsx_bytes)
                st.success(f"완료 — 전 챕터 반복 표현 {n}종")
            except Exception as e:
                st.error(f"오류: {e}")
                import traceback; st.code(traceback.format_exc())

    if 's2_xlsx' in st.session_state:
        st.success(f"전 챕터 반복 패턴 {st.session_state.s2_count}종 → Excel 준비됨")
        st.download_button(
            "⬇ Excel 다운로드 (가나다 순 · 유의어 포함)",
            st.session_state.s2_xlsx,
            f"STEP2_반복표현Excel_{src_file}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        if IS_LOCAL and report_dir:
            st.caption(f"💾 자동 저장됨: `STEP2_반복표현Excel_{src_file}.xlsx`")


# ════════════════════════════════════════════════════════════
# STEP 3 — 직급·계급 체크
# ════════════════════════════════════════════════════════════

_step3_label = "STEP 3 — 직급·계급 역행 체크  |  무료 · 소설 전용" if is_novel else "STEP 3 — 직급·계급 역행 체크  |  소설 전용 (현재 장르에는 불필요)"
with st.expander(_step3_label):
    st.markdown("""
    군계급(이병→일병→상병→병장) 또는 직장직급(사원→대리→과장→부장)이
    챕터 진행상 역행하는 경우를 찾습니다.
    계급 체계가 없는 소설은 이 STEP을 건너뛰세요.
    """)

    ranks_input = st.text_input(
        "계급 순서 (낮은 → 높은, 쉼표 구분)",
        placeholder="예: 이병,일병,상병,병장  또는  사원,대리,과장,부장,사장",
        key="ranks"
    )
    if st.button("STEP 3 실행", key="r3", type="primary"):
        if not ranks_input.strip():
            st.warning("계급 순서를 입력하거나, 없으면 건너뛰세요.")
        else:
            ranks = [r.strip() for r in ranks_input.split(',') if r.strip()]
            with st.spinner("직급 체크 중..."):
                try:
                    r = run_step3(chapters, ranks)
                    st.session_state.s3 = r
                    n = len(r['regressions'])
                    st.success(f"완료 — {'역행 의심 ' + str(n) + '건' if n else '이상 없음'}")
                except Exception as e:
                    st.error(f"오류: {e}")

    if 's3' in st.session_state:
        import pandas as pd
        r = st.session_state.s3
        st.markdown("---")
        if r['regressions']:
            st.markdown('<div class="warn">아래 항목을 직접 확인하세요. 회상 장면이나 동명이인일 수 있습니다.</div>',
                        unsafe_allow_html=True)
            for iss in r['regressions']:
                st.error(f"**{iss['name']}**: {iss['prev_rank']}({iss['prev_ch']}화) → "
                         f"{iss['curr_rank']}({iss['curr_ch']}화)  _{iss.get('sentence','')[:80]}_")
        else:
            st.markdown('<div class="ok">계급 역행 없음 — 이상 없습니다</div>', unsafe_allow_html=True)

        if r.get('timeline'):
            rows = [{'인물':name,'챕터':f'{ch}화','계급':rank}
                    for name, refs in r['timeline'].items()
                    for ch,rank,_ in refs]
            if rows:
                st.markdown("**인물별 계급 타임라인**")
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        json_bytes = json.dumps(r, ensure_ascii=False, indent=2).encode('utf-8')
        st.download_button("⬇ STEP3 계급체크 결과 JSON 저장", json_bytes,
                           f"STEP3_계급체크_{src_file}.json", "application/json")
        save_local(report_dir, f"STEP3_계급체크_{src_file}.json", json_bytes)


# ════════════════════════════════════════════════════════════
# STEP 4 — 페이싱 차트 (2000px 이상, A4 비율)
# ════════════════════════════════════════════════════════════

with st.expander("STEP 4 — 페이싱 차트  |  무료 · 챕터별 글자수·대화비중 그래프"):
    st.markdown("""
    챕터별 3가지 지표를 **큰 그래프(2000px 이상, A4 비율)**로 시각화합니다.
    각 꼭지점에 "N화-수치" 라벨이 표시됩니다.
    평균선(회색 점선)에서 크게 벗어난 챕터를 검토해보세요.
    """)

    if st.button("STEP 4 실행", key="r4", type="primary"):
        with st.spinner("차트 생성 중..."):
            try:
                png_bytes, rows = run_step4_chart(chapters)
                if png_bytes:
                    st.session_state.s4_png  = png_bytes
                    st.session_state.s4_rows = rows
                    save_local(report_dir, f"STEP4_페이싱차트_{src_file}.png", png_bytes)
                    st.success(f"완료 — {len(rows)}개 챕터")
            except Exception as e:
                st.error(f"오류: {e}")
                import traceback; st.code(traceback.format_exc())

    if 's4_png' in st.session_state:
        import pandas as pd
        st.image(st.session_state.s4_png, use_container_width=True)
        st.download_button("⬇ 차트 PNG 다운로드 (고해상도)",
                           st.session_state.s4_png,
                           f"STEP4_페이싱차트_{src_file}.png", "image/png")
        if IS_LOCAL and report_dir:
            st.caption(f"💾 자동 저장됨: `STEP4_페이싱차트_{src_file}.png`")

        rows = st.session_state.s4_rows
        df = pd.DataFrame([{'챕터':r['label'],'글자수':f"{r['chars']:,}",
                             '문장평균길이':f"{r['sent_mean']:.1f}자",
                             '대화비중':f"{r['dia']:.1f}%"} for r in rows])
        st.dataframe(df, use_container_width=True, hide_index=True)


# ════════════════════════════════════════════════════════════
# STEP 5 — 인물 등장 추적 (이름 직접 입력)
# ════════════════════════════════════════════════════════════

_step5_label = "STEP 5 — 인물 등장 추적  |  무료 · 소설 전용" if is_novel else "STEP 5 — 인물 등장 추적  |  소설 전용 (현재 장르에는 불필요)"
with st.expander(_step5_label):
    st.markdown("""
    분석하고 싶은 **인물 이름을 직접 입력**하면,
    각 인물이 어느 챕터에 몇 번 등장했는지, 두 인물이 같이 등장한 챕터는 어디인지 보여줍니다.

    예: `강민준, 서하은, 이대리, 최부장`
    """)

    char_input = st.text_input("인물 이름 (쉼표 또는 공백으로 구분)",
                               placeholder="예: 강민준, 서하은, 이대리",
                               key="char_names")

    if st.button("STEP 5 실행", key="r5", type="primary"):
        if not char_input.strip():
            st.warning("인물 이름을 입력해주세요.")
        else:
            with st.spinner("인물 추적 중..."):
                try:
                    result = run_step5_characters(chapters, char_input)
                    if result and result[0]:
                        st.session_state.s5 = result
                        data, all_labels, co = result
                        st.success(f"완료 — {len(data)}명 추적")
                    else:
                        st.warning("이름을 찾을 수 없습니다. 소설에 실제로 등장하는 이름인지 확인하세요.")
                except Exception as e:
                    st.error(f"오류: {e}")

    if 's5' in st.session_state:
        import pandas as pd
        data, all_labels, co = st.session_state.s5
        names = list(data.keys())

        st.markdown("---")
        st.markdown("**챕터별 등장 횟수**")
        # 인물 카드
        cols = st.columns(min(len(names), 5))
        for i, name in enumerate(names):
            total = sum(data[name].values())
            ch_cnt = len(data[name])
            with cols[i % len(cols)]:
                st.markdown(f"""
                <div class="char-card">
                  <div class="char-name">{name}</div>
                  <div class="char-cnt">{total}회</div>
                  <div style="font-size:.72rem;color:#888">{ch_cnt}개 챕터</div>
                </div>""", unsafe_allow_html=True)

        # 챕터 × 인물 매트릭스
        matrix = []
        for ch in all_labels:
            row = {'챕터': ch}
            for name in names:
                row[name] = data[name].get(ch, 0)
            matrix.append(row)
        df = pd.DataFrame(matrix)
        st.dataframe(df, use_container_width=True, hide_index=True)

        # 공동 등장 관계
        if co:
            st.markdown("**공동 등장 챕터 수** (두 인물이 같은 챕터에 나오는 경우)")
            co_rows = [{'인물 A': a, '인물 B': b, '공동 등장 챕터 수': n}
                       for (a,b),n in sorted(co.items(), key=lambda x:-x[1]) if n > 0]
            if co_rows:
                st.dataframe(pd.DataFrame(co_rows), use_container_width=True, hide_index=True)

        # 저장
        json_bytes = json.dumps({'characters': {n: data[n] for n in names}},
                                ensure_ascii=False, indent=2).encode('utf-8')
        st.download_button("⬇ 인물 추적 결과 JSON 저장", json_bytes,
                           f"STEP5_인물추적_{src_file}.json", "application/json")
        save_local(report_dir, f"STEP5_인물추적_{src_file}.json", json_bytes)


st.markdown("---")
st.markdown("""
<div class="step-section-header step-section-header-gold">
  <div>
    <div class="step-section-title">STEP 6 – 8 &nbsp; AI 분석 (Claude.ai 무료 사용)</div>
    <div class="step-section-desc">
      Claude Pro 구독만 있으면 API 키 없이 사용 가능 &nbsp;·&nbsp;
      프롬프트 생성 → claude.ai 붙여넣기 → 결과 저장
    </div>
  </div>
</div>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# STEP 6 무료 — 프롬프트 생성 → Claude.ai 붙여넣기
# ════════════════════════════════════════════════════════════

with st.expander("STEP 6 — AI 반복 표현 진단  |  Claude.ai 붙여넣기 · 무료 (Pro 구독만 있으면 OK)"):
    st.markdown("""
    STEP 1~2에서 찾은 반복 표현 목록을 **프롬프트로 만들어** claude.ai 채팅에 붙여넣으면
    진단 + 대안 표현 3개를 바로 받을 수 있습니다.

    **Claude Pro 구독만 있으면 추가 비용 없이 사용 가능합니다.**

    사용 순서: ① 아래 버튼 클릭 → ② 생성된 텍스트 복사 →
    ③ [claude.ai](https://claude.ai) 채팅창에 붙여넣기 → ④ 결과를 아래 칸에 저장
    """)

    if st.button("STEP 6 프롬프트 생성", key="r6", type="primary"):
        with st.spinner("반복 표현 목록 수집 중..."):
            try:
                prompt6, n6 = build_step6_prompt(chapters)
                if prompt6:
                    st.session_state.s6_prompt = prompt6
                    st.session_state.s6_n = n6
                    st.success(f"완료 — 반복 표현 {n6}종 포함")
                else:
                    st.info("반복 표현이 발견되지 않았습니다. STEP 1을 먼저 실행해 보세요.")
            except Exception as e:
                st.error(f"오류: {e}")

    if 's6_prompt' in st.session_state:
        st.markdown(f"**아래 텍스트를 복사 → [claude.ai](https://claude.ai) 채팅에 붙여넣기:**")
        st.code(st.session_state.s6_prompt, language=None)
        st.download_button("⬇ 프롬프트 TXT 저장", st.session_state.s6_prompt.encode('utf-8'),
                           f"STEP6_프롬프트_{src_file}.txt", "text/plain", key="dl6p")
        save_local(report_dir, f"STEP6_프롬프트_{src_file}.txt",
                   st.session_state.s6_prompt.encode('utf-8'))
        st.markdown("---")
        st.markdown("**Claude 결과 붙여넣기** (선택 — 로컬 저장용)")
        s6_paste = st.text_area("claude.ai 답변을 여기 붙여넣으면 자동 저장됩니다",
                                 height=150, key="s6_paste",
                                 placeholder="claude.ai 답변을 여기에 붙여넣으세요...")
        if s6_paste.strip():
            save_local(report_dir, f"STEP6_AI진단_{src_file}.md", s6_paste.encode('utf-8'))
            st.download_button("⬇ 결과 MD 저장", s6_paste.encode('utf-8'),
                               f"STEP6_AI진단_{src_file}.md", "text/markdown", key="dl6r")
            if IS_LOCAL and report_dir:
                st.caption(f"💾 자동 저장됨: `{report_dir}`")


# ════════════════════════════════════════════════════════════
# STEP 6+ 유료 — API 자동화
# ════════════════════════════════════════════════════════════

with st.expander("STEP 6+ — AI 진단 자동화  |  API 키 보유자 전용 · 앱 안에서 즉시 처리"):
    st.markdown("""
    API 키가 있다면 반복 표현을 **앱 안에서 자동으로 처리**합니다.
    claude.ai에 붙여넣는 과정 없이 바로 결과 표시 + Excel 저장까지 한번에.

    - 모델: `claude-haiku` | 비용: 표현 20개 기준 약 $0.01~$0.05
    """)

    if not st.session_state.get('api_key'):
        st.markdown('<div class="warn">사이드바에 Claude API 키를 입력해야 실행할 수 있습니다.</div>',
                    unsafe_allow_html=True)
    else:
        if st.button("STEP 6+ 실행 (API 자동화)", key="r6p", type="primary"):
            with st.spinner("Claude AI 분석 중..."):
                try:
                    result = run_step6_ai_suggest(chapters, st.session_state.api_key)
                    st.session_state.s6p = result
                    st.success(f"완료 — {len(result)}개 표현 진단")
                except Exception as e:
                    st.error(f"오류: {e}")

    if 's6p' in st.session_state:
        import pandas as pd
        result = st.session_state.s6p
        rows = [{'표현': p, '횟수': info['count'], 'AI 진단 + 대안': info['response']}
                for p, info in result.items()]
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        try:
            import openpyxl
            from openpyxl.styles import Alignment
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'AI진단'
            ws.append(['표현', '횟수', 'AI 진단 + 대안', '실제 문장 예시'])
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 60
            for p, info in result.items():
                ctx = ' | '.join(info.get('contexts', [])[:2])
                ws.append([p, info['count'], info['response'], ctx])
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            buf = io.BytesIO(); wb.save(buf)
            xlsx_bytes = buf.getvalue()
            st.download_button("⬇ AI 진단 Excel 저장", xlsx_bytes,
                               f"AI진단_{proj_name}.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl6px")
            save_local(report_dir, f"STEP6_AI진단_{src_file}.xlsx", xlsx_bytes)
        except Exception:
            pass


st.markdown("")


# ════════════════════════════════════════════════════════════
# STEP 7 무료 — 복선 탐지 + 회수 확인 (프롬프트 생성)
# ════════════════════════════════════════════════════════════

_step7_label = "STEP 7 — 복선 탐지 + 회수 확인  |  소설 전용 · Claude.ai 붙여넣기 · 무료" if is_novel else "STEP 7 — 복선 탐지 + 회수 확인  |  소설 전용 (비소설에는 불필요)"
with st.expander(_step7_label):
    st.markdown("""
    **두 단계로 진행합니다.** Claude Pro 구독만 있으면 추가 비용 없이 사용 가능합니다.

    - **1단계**: 전반부 화수 범위 선택 → 프롬프트 생성 → claude.ai에 붙여넣기 → 복선 목록 받기
    - **2단계**: 1단계 결과를 아래에 붙여넣기 → 후반부 프롬프트 생성 → claude.ai에 붙여넣기 → 회수 여부 판정

    > **팁**: 분량이 많으면 프롬프트 TXT를 다운로드해서 claude.ai **파일 업로드**로 전송하세요.
    """)
    st.warning("⚠️ **STEP 7은 장편소설(픽션)에만 유효합니다.** 비즈니스·에세이·논픽션 책은 복선/회수 구조가 없으므로 이 단계를 건너뛰세요. 해당 장르는 STEP 6(반복 표현) + STEP 8(출판 자료)만 사용하면 충분합니다.")

    _s7_single = _is_single_file(chapters)
    ch_nums_all = sorted(ch_sort(k) for k in chapters if 0 < ch_sort(k) <= 9999)
    first_ch = ch_nums_all[0] if ch_nums_all else 1
    last_ch  = ch_nums_all[-1] if ch_nums_all else 36
    mid_ch   = ch_nums_all[len(ch_nums_all)//2 - 1] if len(ch_nums_all) >= 2 else (last_ch + first_ch) // 2

    st.markdown("### 1단계 — 전반부 떡밥 탐지")
    if _s7_single:
        st.info("챕터 구분(제1화~)이 없는 파일입니다. 전체 본문을 자동으로 전반/후반으로 나눠 분석합니다.")
        s7_s1, s7_e1 = 0, 0
    else:
        c1, c2 = st.columns(2)
        s7_s1 = c1.number_input("전반부 시작 화", min_value=1, value=first_ch, key="s7s1")
        s7_e1 = c2.number_input("전반부 끝 화",   min_value=1, value=mid_ch,   key="s7e1")
        sel1_chars = sum(len(v) for k, v in chapters.items() if s7_s1 <= ch_sort(k) <= s7_e1)
        st.caption(f"선택 범위 약 {sel1_chars:,}자")

    if st.button("1단계 프롬프트 생성", key="r7a", type="primary"):
        with st.spinner("프롬프트 생성 중..."):
            p7a = build_step7_phase1_prompt(chapters, int(s7_s1), int(s7_e1))
            if p7a:
                st.session_state.s7_p1 = p7a
                st.session_state.s7_p1_range = (int(s7_s1), int(s7_e1))
                st.success(f"완료 — {len(p7a):,}자 프롬프트")
            else:
                st.warning(f"{s7_s1}~{s7_e1}화 범위에 챕터가 없습니다.")

    if 's7_p1' in st.session_state:
        r1, r2 = st.session_state.s7_p1_range
        p1 = st.session_state.s7_p1
        fn1 = f"STEP7_1단계_복선탐지_{src_file}_{'전반부' if _s7_single else f'{r1}-{r2}화'}.txt"
        st.download_button(f"⬇ 1단계 프롬프트 TXT 다운로드 ({len(p1):,}자)",
                           p1.encode('utf-8'), fn1, "text/plain", key="dl7p1")
        save_local(report_dir, fn1, p1.encode('utf-8'))
        st.markdown("**[claude.ai 열기](https://claude.ai)** → 파일 업로드 또는 내용 붙여넣기 → 결과 받기")

        st.markdown("---")
        st.markdown("### 2단계 — 후반부 회수 확인")
        st.markdown("**1단계에서 받은 claude.ai 답변을 아래에 붙여넣으세요:**")
        s7_phase1_paste = st.text_area(
            "1단계 결과 붙여넣기",
            height=180, key="s7_p1_paste",
            placeholder="claude.ai에서 받은 복선 목록을 여기에 붙여넣으세요..."
        )
        if _s7_single:
            s7_s2, s7_e2 = 0, 0
        else:
            c3, c4 = st.columns(2)
            s7_s2 = c3.number_input("후반부 시작 화", min_value=1, value=mid_ch + 1, key="s7s2")
            s7_e2 = c4.number_input("후반부 끝 화",   min_value=1, value=last_ch,    key="s7e2")
            sel2_chars = sum(len(v) for k, v in chapters.items() if s7_s2 <= ch_sort(k) <= s7_e2)
            st.caption(f"선택 범위 약 {sel2_chars:,}자")

        if st.button("2단계 프롬프트 생성", key="r7b", type="primary"):
            if not s7_phase1_paste.strip():
                st.warning("1단계 claude.ai 답변을 먼저 붙여넣으세요.")
            else:
                with st.spinner("프롬프트 생성 중..."):
                    p7b = build_step7_phase2_prompt(
                        s7_phase1_paste, chapters, int(s7_s2), int(s7_e2))
                    if p7b:
                        st.session_state.s7_p2 = p7b
                        st.session_state.s7_p2_range = (int(s7_s2), int(s7_e2))
                        save_local(report_dir,
                                   f"STEP7_복선탐지결과_{src_file}.md",
                                   s7_phase1_paste.encode('utf-8'))
                        st.success(f"완료 — {len(p7b):,}자 프롬프트")
                    else:
                        st.warning(f"{s7_s2}~{s7_e2}화 범위에 챕터가 없습니다.")

        if 's7_p2' in st.session_state:
            r3, r4 = st.session_state.s7_p2_range
            p2 = st.session_state.s7_p2
            fn2 = f"STEP7_2단계_회수확인_{src_file}_{'후반부' if _s7_single else f'{r3}-{r4}화'}.txt"
            st.download_button(f"⬇ 2단계 프롬프트 TXT 다운로드 ({len(p2):,}자)",
                               p2.encode('utf-8'), fn2, "text/plain", key="dl7p2")
            save_local(report_dir, fn2, p2.encode('utf-8'))
            st.markdown("**[claude.ai 열기](https://claude.ai)** → 파일 업로드 또는 내용 붙여넣기 → 결과 받기")

            st.markdown("---")
            st.markdown("**2단계 결과 저장** (선택)")
            s7_paste2 = st.text_area("2단계 claude.ai 답변을 여기 붙여넣으면 자동 저장됩니다",
                                      height=150, key="s7_paste2",
                                      placeholder="회수 확인 결과를 붙여넣으세요...")
            if s7_paste2.strip():
                fn_r2 = f"STEP7_복선회수확인결과_{src_file}.md"
                save_local(report_dir, fn_r2, s7_paste2.encode('utf-8'))
                st.download_button("⬇ 결과 MD 저장", s7_paste2.encode('utf-8'),
                                   fn_r2, "text/markdown", key="dl7r2")
                if IS_LOCAL and report_dir:
                    st.caption(f"💾 자동 저장됨: `{report_dir}`")


# ════════════════════════════════════════════════════════════
# STEP 8 무료 — 출판 자료 (프롬프트 생성)
# ════════════════════════════════════════════════════════════

with st.expander("STEP 8 — 출판 자료 생성  |  Claude.ai 붙여넣기 · 무료 (Pro 구독만 있으면 OK)"):
    st.markdown("""
    소설 앞뒤 챕터 샘플로 프롬프트를 만들어 claude.ai에 붙여넣으면
    출판사 제출용 자료를 바로 받을 수 있습니다.

    - 책 소개글 (400자 이내)
    - 줄거리 (600자 이내)
    - 주요 인물 소개
    - 홍보문구 3가지 (각 30자 이내)

    **Claude Pro 구독만 있으면 추가 비용 없이 사용 가능합니다.**
    """)

    pub_title = st.text_input("소설 제목 (출판 자료에 표시)", value=proj_name, key="pub_title")

    if st.button("STEP 8 프롬프트 생성", key="r8", type="primary"):
        with st.spinner("프롬프트 생성 중..."):
            p8 = build_step8_prompt(chapters, pub_title)
            st.session_state.s8_prompt = p8
            st.success("완료")

    if 's8_prompt' in st.session_state:
        p8 = st.session_state.s8_prompt
        st.markdown("**아래 텍스트를 복사 → [claude.ai](https://claude.ai) 채팅에 붙여넣기:**")
        st.code(p8, language=None)
        st.download_button("⬇ 프롬프트 TXT 저장", p8.encode('utf-8'),
                           f"STEP8_프롬프트_{src_file}.txt", "text/plain", key="dl8p")
        save_local(report_dir, f"STEP8_프롬프트_{src_file}.txt", p8.encode('utf-8'))
        st.markdown("---")
        st.markdown("**결과 저장** (선택)")
        s8_paste = st.text_area("claude.ai 답변을 여기 붙여넣으면 자동 저장됩니다",
                                 height=150, key="s8_paste",
                                 placeholder="출판 자료 결과를 붙여넣으세요...")
        if s8_paste.strip():
            save_local(report_dir, f"STEP8_출판자료_{src_file}.md", s8_paste.encode('utf-8'))
            st.download_button("⬇ 출판 자료 MD 저장", s8_paste.encode('utf-8'),
                               f"STEP8_출판자료_{src_file}.md", "text/markdown", key="dl8r")
            if IS_LOCAL and report_dir:
                st.caption(f"💾 자동 저장됨: `STEP8_출판자료_{src_file}.md`")

st.markdown("---")
st.markdown("""
<div style="text-align:center;font-size:.72rem;color:#BBB;padding:16px;line-height:2">
  NOVELDESK &nbsp;·&nbsp; AI 소설 분석 도구<br>
  Claude Pro 구독만 있으면 API 키 없이 전체 기능 사용 가능
</div>
""", unsafe_allow_html=True)
